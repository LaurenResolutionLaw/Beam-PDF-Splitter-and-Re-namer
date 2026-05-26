from __future__ import annotations

import hashlib
import importlib.util
import os
import re
import shutil
import sys
import zipfile
from dataclasses import asdict, dataclass
from io import BytesIO
from pathlib import Path
from typing import Callable

import fitz
import streamlit as st


APP_TITLE = "Resolution Law Tools"
FOOTER_TOP_RATIO = 0.92
INVALID_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


@dataclass
class ToolDefinition:
    tool_id: str
    name: str
    category: str
    description: str
    render: Callable[[], None]


@dataclass
class ChunkRow:
    row_id: str
    source_file: str
    pages: str
    start_page: int
    end_page: int
    detected_beam: str
    filename_stem: str
    method: str
    issue: str
    include: bool


@dataclass
class SplitResult:
    source_file: str
    pages: str
    output_file: str
    beam_number: str
    method: str
    status: str


def inject_css() -> None:
    st.markdown(
        """
        <style>
        .stApp {
            background:
                linear-gradient(180deg, #f8fafc 0%, #eef4ff 44%, #f8fafc 100%);
            color: #0f172a;
        }
        [data-testid="stHeader"] {
            background: rgba(248, 250, 252, 0.88);
        }
        [data-testid="stSidebar"] {
            background: #0f172a;
        }
        [data-testid="stSidebar"] * {
            color: #e5e7eb;
        }
        [data-testid="stSidebar"] div[role="radiogroup"] label {
            border-radius: 12px;
            padding: 0.25rem 0.35rem;
        }
        .hub-hero {
            background: #ffffff;
            color: #0f172a;
            padding: 2rem 2.15rem;
            border: 1px solid #dbe4f0;
            border-radius: 20px;
            margin-bottom: 1.25rem;
            box-shadow: 0 18px 45px rgba(15, 23, 42, 0.08);
            position: relative;
            overflow: hidden;
        }
        .hub-hero:before {
            content: "";
            position: absolute;
            inset: 0 0 auto 0;
            height: 5px;
            background: linear-gradient(90deg, #2563eb, #14b8a6, #facc15);
        }
        .hub-hero h1 {
            margin: 0;
            font-size: 2.35rem;
            line-height: 1.05;
            letter-spacing: 0;
            color: #0f172a;
        }
        .hub-hero p {
            margin: 0.75rem 0 0;
            color: #475569;
            font-size: 1.03rem;
            max-width: 760px;
        }
        .tool-card {
            background: white;
            border: 1px solid #e2e8f0;
            border-radius: 16px;
            padding: 1.15rem 1.2rem;
            box-shadow: 0 10px 28px rgba(15, 23, 42, 0.05);
            min-height: 134px;
            transition: border-color 140ms ease, box-shadow 140ms ease, transform 140ms ease;
        }
        .tool-card.clickable {
            min-height: 160px;
        }
        .tool-card.clickable:hover {
            border-color: #93c5fd;
            box-shadow: 0 18px 40px rgba(37, 99, 235, 0.12);
            transform: translateY(-1px);
        }
        .tool-card h3 {
            margin: 0 0 0.35rem;
            color: #0f172a;
        }
        .tool-card p,
        .small-muted {
            color: #667085;
            font-size: 0.92rem;
        }
        .metric-strip {
            display: grid;
            grid-template-columns: repeat(3, minmax(0, 1fr));
            gap: 0.75rem;
            margin: 1rem 0;
        }
        .metric-box {
            background: white;
            border: 1px solid #e2e8f0;
            border-radius: 14px;
            padding: 0.85rem 1rem;
            box-shadow: 0 6px 18px rgba(15, 23, 42, 0.04);
        }
        .metric-box strong {
            display: block;
            color: #0f172a;
            font-size: 1.45rem;
        }
        .metric-box span {
            color: #667085;
            font-size: 0.88rem;
        }
        div[data-testid="stDownloadButton"] button,
        div[data-testid="stButton"] button {
            border-radius: 12px;
            font-weight: 650;
            border: 1px solid #cbd5e1;
            box-shadow: 0 5px 14px rgba(15, 23, 42, 0.06);
        }
        .section-label {
            color: #0f172a;
            font-size: 1.1rem;
            font-weight: 700;
            margin: 1rem 0 0.25rem;
        }
        .stTabs [data-baseweb="tab-list"] {
            gap: 0.5rem;
        }
        .stTabs [data-baseweb="tab"] {
            border-radius: 999px;
            background: #ffffff;
            border: 1px solid #e2e8f0;
            padding: 0.45rem 0.85rem;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def find_tesseract_executable() -> Path | None:
    env_value = os.environ.get("TESSERACT_CMD")
    if env_value and Path(env_value).exists():
        return Path(env_value)

    path_value = shutil.which("tesseract")
    if path_value:
        return Path(path_value)

    if sys.platform.startswith("win"):
        candidates = [
            Path(os.environ.get("ProgramFiles", r"C:\Program Files")) / "Tesseract-OCR" / "tesseract.exe",
            Path(os.environ.get("ProgramFiles(x86)", r"C:\Program Files (x86)")) / "Tesseract-OCR" / "tesseract.exe",
        ]
        for candidate in candidates:
            if candidate.exists():
                return candidate

    return None


def dependency_status() -> dict[str, bool]:
    return {
        "PyMuPDF": importlib.util.find_spec("fitz") is not None,
        "Pillow": importlib.util.find_spec("PIL") is not None,
        "pytesseract": importlib.util.find_spec("pytesseract") is not None,
        "Tesseract OCR": find_tesseract_executable() is not None,
    }


def ocr_available() -> bool:
    return (
        importlib.util.find_spec("pytesseract") is not None
        and importlib.util.find_spec("PIL") is not None
        and find_tesseract_executable() is not None
    )


def configure_tesseract(pytesseract_module) -> None:
    executable = find_tesseract_executable()
    if executable is not None:
        pytesseract_module.pytesseract.tesseract_cmd = str(executable)


def sanitize_filename_stem(value: str, fallback: str) -> str:
    stem = value.strip()
    if stem.lower().endswith(".pdf"):
        stem = stem[:-4]
    stem = INVALID_FILENAME_CHARS.sub("_", stem)
    stem = re.sub(r"\s+", "_", stem).strip(" ._")
    return stem or fallback


def make_unique_stem(existing: set[str], requested_stem: str, fallback: str) -> str:
    base = sanitize_filename_stem(requested_stem, fallback)
    candidate = base
    counter = 2
    while candidate.lower() in existing:
        candidate = f"{base}_{counter:02d}"
        counter += 1
    existing.add(candidate.lower())
    return candidate


def number_tokens(text: str) -> list[str]:
    if not text:
        return []

    normalized = text.replace(",", "")
    tokens = re.findall(r"\b\d{2,}\b", normalized)
    if tokens:
        return tokens

    digits_only = re.sub(r"\D", "", normalized)
    if len(digits_only) >= 2:
        return [digits_only]

    return []


def find_beam_number_in_text(text: str) -> str | None:
    tokens = number_tokens(text)
    return tokens[-1] if tokens else None


def find_native_footer_number(page) -> str | None:
    page_rect = page.rect
    footer_y = page_rect.height * 0.74
    candidates: list[tuple[float, float, str]] = []

    for word in page.get_text("words"):
        x0, y0, x1, y1, text = word[:5]
        if y0 < footer_y:
            continue
        for token in number_tokens(text):
            candidates.append((float(y1), float(x0), token))

    if not candidates:
        text = page.get_text("text", clip=fitz.Rect(0, footer_y, page_rect.width, page_rect.height)) or ""
        return find_beam_number_in_text(text)

    candidates.sort(key=lambda item: (item[0], item[1]))
    return candidates[-1][2]


def find_ocr_footer_number(page, ocr_error_callback: Callable[[Exception], None] | None = None) -> tuple[str | None, str]:
    try:
        from PIL import Image, ImageOps
        import pytesseract

        configure_tesseract(pytesseract)
        page_rect = page.rect
        footer_rect = fitz.Rect(0, page_rect.height * 0.74, page_rect.width, page_rect.height)
        matrix = fitz.Matrix(3, 3)
        pixmap = page.get_pixmap(matrix=matrix, clip=footer_rect, alpha=False)
        image = Image.frombytes("RGB", [pixmap.width, pixmap.height], pixmap.samples)
        image = ImageOps.autocontrast(ImageOps.grayscale(image))

        data = pytesseract.image_to_data(
            image,
            config="--psm 6 -c tessedit_char_whitelist=0123456789",
            output_type=pytesseract.Output.DICT,
        )

        candidates: list[tuple[float, float, str]] = []
        for index, text in enumerate(data.get("text", [])):
            for token in number_tokens(text):
                top = float(data["top"][index])
                height = float(data["height"][index])
                left = float(data["left"][index])
                candidates.append((top + height, left, token))

        if candidates:
            candidates.sort(key=lambda item: (item[0], item[1]))
            return candidates[-1][2], "OCR"

        ocr_text = pytesseract.image_to_string(
            image,
            config="--psm 6 -c tessedit_char_whitelist=0123456789",
        )
        beam = find_beam_number_in_text(ocr_text)
        if beam:
            return beam, "OCR"
    except Exception as exc:
        if ocr_error_callback is not None:
            ocr_error_callback(exc)

    return None, "not found"


def extract_beam_number_from_page(
    page,
    use_ocr: bool,
    ocr_error_callback: Callable[[Exception], None] | None = None,
) -> tuple[str | None, str]:
    beam = find_native_footer_number(page)
    if beam:
        return beam, "native text"

    if not use_ocr:
        return None, "not found"

    return find_ocr_footer_number(page, ocr_error_callback)


def save_page_range_to_bytes(doc, start_page: int, end_page: int) -> bytes:
    out_doc = fitz.open()
    try:
        out_doc.insert_pdf(doc, from_page=start_page, to_page=end_page)
        buffer = BytesIO()
        out_doc.save(buffer, garbage=4, deflate=True)
        return buffer.getvalue()
    finally:
        out_doc.close()


def file_hash(data: bytes) -> str:
    return hashlib.sha1(data).hexdigest()[:12]


def uploaded_files_signature(uploaded_files) -> str:
    parts: list[str] = []
    for uploaded_file in uploaded_files:
        data = uploaded_file.getvalue()
        parts.append(f"{uploaded_file.name}:{len(data)}:{file_hash(data)}")
    return "|".join(sorted(parts))


def analyze_uploads(uploaded_files, odd_page_mode: str, missing_beam_mode: str) -> tuple[list[dict], list[str], dict[str, bytes]]:
    rows: list[ChunkRow] = []
    logs: list[str] = []
    file_bytes: dict[str, bytes] = {}
    fallback_counter = 1
    use_ocr = ocr_available()
    ocr_warning_logged = False

    if not use_ocr:
        logs.append("OCR is unavailable. Native PDFs can still be processed, but scanned PDFs need Tesseract OCR on the server.")

    def warn_ocr_once(exc: Exception) -> None:
        nonlocal ocr_warning_logged
        if not ocr_warning_logged:
            logs.append(f"OCR warning: {exc}")
            ocr_warning_logged = True

    for uploaded_file in uploaded_files:
        data = uploaded_file.getvalue()
        source_name = Path(uploaded_file.name).name
        source_id = f"{file_hash(data)}_{source_name}"
        file_bytes[source_id] = data

        try:
            doc = fitz.open(stream=data, filetype="pdf")
        except Exception as exc:
            logs.append(f"{source_name}: could not open PDF ({exc}).")
            continue

        try:
            page_count = doc.page_count
            logs.append(f"{source_name}: opened {page_count} page(s).")
            paired_page_count = page_count - (page_count % 2)

            for start_page in range(0, paired_page_count, 2):
                beam_number, method = extract_beam_number_from_page(doc[start_page + 1], use_ocr, warn_ocr_once)
                issue = ""
                include = True
                filename_stem = beam_number or ""

                if not beam_number:
                    if missing_beam_mode == "Skip missing beam pairs":
                        include = False
                        issue = "No beam number found; skipped by rule."
                    elif missing_beam_mode == "Use fallback names":
                        filename_stem = f"unnamed_{fallback_counter:03d}"
                        fallback_counter += 1
                        issue = "No beam number found; fallback name assigned."
                    else:
                        issue = "No beam number found; enter filename or uncheck include."

                rows.append(
                    ChunkRow(
                        row_id=f"{source_id}:{start_page}:{start_page + 1}",
                        source_file=source_name,
                        pages=f"{start_page + 1}-{start_page + 2}",
                        start_page=start_page,
                        end_page=start_page + 1,
                        detected_beam=beam_number or "",
                        filename_stem=filename_stem,
                        method=method,
                        issue=issue,
                        include=include,
                    )
                )

            if page_count % 2:
                orphan_page = page_count - 1
                if odd_page_mode == "Skip odd final pages":
                    filename_stem = ""
                    issue = "Odd final page; skipped by rule."
                    include = False
                elif odd_page_mode == "Review odd final pages":
                    filename_stem = ""
                    issue = "Odd final page; enter filename or uncheck include."
                    include = True
                else:
                    filename_stem = f"{Path(source_name).stem}_page_{page_count:03d}"
                    issue = "Odd final page saved as single-page output."
                    include = True

                rows.append(
                    ChunkRow(
                        row_id=f"{source_id}:{orphan_page}:{orphan_page}",
                        source_file=source_name,
                        pages=str(page_count),
                        start_page=orphan_page,
                        end_page=orphan_page,
                        detected_beam="",
                        filename_stem=filename_stem,
                        method="odd page",
                        issue=issue,
                        include=include,
                    )
                )
        finally:
            doc.close()

    return [asdict(row) for row in rows], logs, file_bytes


def build_zip(rows: list[dict], file_bytes: dict[str, bytes]) -> tuple[bytes | None, list[str]]:
    errors: list[str] = []
    output = BytesIO()
    used_names: set[str] = set()
    written = 0

    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for row in rows:
            if not row.get("include"):
                continue

            filename_stem = str(row.get("filename_stem") or "").strip()
            if not filename_stem:
                errors.append(f"{row['source_file']} pages {row['pages']}: filename is blank.")
                continue

            source_key = str(row["row_id"]).split(":", 1)[0]
            data = file_bytes.get(source_key)
            if data is None:
                errors.append(f"{row['source_file']} pages {row['pages']}: source file data was not found.")
                continue

            try:
                doc = fitz.open(stream=data, filetype="pdf")
                try:
                    pdf_bytes = save_page_range_to_bytes(doc, int(row["start_page"]), int(row["end_page"]))
                finally:
                    doc.close()
            except Exception as exc:
                errors.append(f"{row['source_file']} pages {row['pages']}: could not split PDF ({exc}).")
                continue

            unique_stem = make_unique_stem(used_names, filename_stem, f"unnamed_{written + 1:03d}")
            archive.writestr(f"beam_split_results/{unique_stem}.pdf", pdf_bytes)
            written += 1

    if errors:
        return None, errors
    if written == 0:
        return None, ["No output files were selected. Check at least one row to include."]

    return output.getvalue(), []


def split_and_rename_pdfs(uploaded_files, include_odd_final_page: bool, use_fallback_names: bool) -> tuple[bytes | None, list[dict], list[str]]:
    logs: list[str] = []
    rows: list[SplitResult] = []
    output = BytesIO()
    used_names: set[str] = set()
    fallback_counter = 1
    written = 0
    use_ocr = ocr_available()
    ocr_warning_logged = False

    if not use_ocr:
        logs.append("OCR is unavailable on the server. Native PDFs may still work, but scanned PDFs need Tesseract OCR.")

    def warn_ocr_once(exc: Exception) -> None:
        nonlocal ocr_warning_logged
        if not ocr_warning_logged:
            logs.append(f"OCR warning: {exc}")
            ocr_warning_logged = True

    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for uploaded_file in uploaded_files:
            data = uploaded_file.getvalue()
            source_name = Path(uploaded_file.name).name

            try:
                doc = fitz.open(stream=data, filetype="pdf")
            except Exception as exc:
                logs.append(f"{source_name}: could not open PDF ({exc}).")
                continue

            try:
                page_count = doc.page_count
                pair_page_count = page_count - (page_count % 2)

                for start_page in range(0, pair_page_count, 2):
                    end_page = start_page + 1
                    pages = f"{start_page + 1}-{end_page + 1}"
                    beam_number, method = extract_beam_number_from_page(doc[end_page], use_ocr, warn_ocr_once)
                    status = "Beam number found."

                    if beam_number:
                        stem = beam_number
                    elif use_fallback_names:
                        stem = f"unnamed_{fallback_counter:03d}"
                        fallback_counter += 1
                        status = "Beam number not found; fallback name used."
                    else:
                        logs.append(f"{source_name} pages {pages}: beam number not found; skipped.")
                        rows.append(
                            SplitResult(
                                source_file=source_name,
                                pages=pages,
                                output_file="",
                                beam_number="",
                                method="not found",
                                status="Skipped - beam number not found.",
                            )
                        )
                        continue

                    unique_stem = make_unique_stem(used_names, stem, f"unnamed_{fallback_counter:03d}")
                    output_file = f"{unique_stem}.pdf"
                    pdf_bytes = save_page_range_to_bytes(doc, start_page, end_page)
                    archive.writestr(output_file, pdf_bytes)
                    written += 1
                    rows.append(
                        SplitResult(
                            source_file=source_name,
                            pages=pages,
                            output_file=output_file,
                            beam_number=beam_number or "",
                            method=method,
                            status=status,
                        )
                    )

                if page_count % 2:
                    last_page = page_count - 1
                    if include_odd_final_page:
                        stem = f"{Path(source_name).stem}_page_{page_count:03d}"
                        unique_stem = make_unique_stem(used_names, stem, stem)
                        output_file = f"{unique_stem}.pdf"
                        pdf_bytes = save_page_range_to_bytes(doc, last_page, last_page)
                        archive.writestr(output_file, pdf_bytes)
                        written += 1
                        rows.append(
                            SplitResult(
                                source_file=source_name,
                                pages=str(page_count),
                                output_file=output_file,
                                beam_number="",
                                method="odd final page",
                                status="Included as single-page output.",
                            )
                        )
                    else:
                        logs.append(f"{source_name} page {page_count}: odd final page skipped.")
            finally:
                doc.close()

    if written == 0:
        return None, [asdict(row) for row in rows], logs + ["No PDF chunks were created."]

    return output.getvalue(), [asdict(row) for row in rows], logs


# ============================================================================
# Spreadsheet Comparison Tool — Resolution Law Tools
# Compares two snapshots of the same tracker spreadsheet (.xlsx or .csv).
# User picks one or more key columns (composite key supported for cases where
# a single column isn't unique per row, e.g. Beam Number + Garnishee when one
# case has multiple garnishments). Reports added rows, removed rows, and per-
# cell changes after normalizing away formatting noise.
# ============================================================================
import csv as _csv
import datetime as _dt
import io as _io
import re as _re
from collections import Counter as _Counter

import openpyxl as _openpyxl
from openpyxl.styles import Font as _Font, PatternFill as _PatternFill, Alignment as _Alignment

_DATE_RE    = _re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{4})$")
_TIME_RE    = _re.compile(r"^(\d{1,2}):(\d{2}):(\d{2})\s*(AM|PM|am|pm)?$")
_MONEY_RE   = _re.compile(r"^\(\s*-?\$?\s*-?[\d,]+(?:\.\d+)?\s*\)$|^-?\$\s*-?[\d,]+(?:\.\d+)?$")
_ILLEGAL_RE = _re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

_KEY_HINTS = [
    "Beam Number", "Beam #", "Beam",
    "Case Number", "Case #", "Case No", "Case",
    "Account ID", "Account_ID", "Client Account ID", "Account",
    "Matter Number", "Matter #", "Matter ID", "Matter",
    "File Number", "File #", "File ID",
    "ID", "Id",
]


def _norm_value(v) -> str:
    if v is None:
        return ""
    if isinstance(v, _dt.datetime):
        return f"{v.month:02d}/{v.day:02d}/{v.year}"
    if isinstance(v, _dt.date):
        return f"{v.month:02d}/{v.day:02d}/{v.year}"
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        if v != v:
            return ""
        if v.is_integer():
            return str(int(v))
        return f"{v:.6f}".rstrip("0").rstrip(".")
    s = str(v).replace(" ", " ")
    s = _ILLEGAL_RE.sub("", s).strip()
    s = _re.sub(r"\s+", " ", s)
    if s == "":
        return ""
    m = _DATE_RE.match(s)
    if m:
        mo, da, yr = m.groups()
        return f"{int(mo):02d}/{int(da):02d}/{yr}"
    m = _TIME_RE.match(s)
    if m:
        hh, mm, ss, ap = m.groups()
        suf = f" {ap.upper()}" if ap else ""
        return f"{int(hh):02d}:{mm}:{ss}{suf}"
    if _MONEY_RE.match(s):
        raw = s.replace(",", "")
        neg = False
        if raw.startswith("(") and raw.endswith(")"):
            neg = True; raw = raw[1:-1].strip()
        if raw.startswith("-"):
            neg = not neg; raw = raw[1:].strip()
        if raw.startswith("$"):
            raw = raw[1:].strip()
        if raw.startswith("-"):
            neg = not neg; raw = raw[1:].strip()
        try:
            val = float(raw)
            if neg: val = -val
            return f"${val:.2f}"
        except ValueError:
            pass
    try:
        f = float(s.replace(",", ""))
        if f.is_integer():
            return str(int(f))
    except (ValueError, TypeError):
        pass
    return s


def _display_value(v) -> str:
    if v is None:
        return ""
    if isinstance(v, _dt.datetime):
        if v.hour == 0 and v.minute == 0 and v.second == 0:
            return f"{v.month:02d}/{v.day:02d}/{v.year}"
        return v.strftime("%m/%d/%Y %H:%M:%S")
    if isinstance(v, _dt.date):
        return f"{v.month:02d}/{v.day:02d}/{v.year}"
    if isinstance(v, float):
        if v != v: return ""
        if v.is_integer(): return str(int(v))
        return f"{v}"
    s = str(v)
    s = _ILLEGAL_RE.sub("", s)
    return s


def _detect_header_row(rows):
    for i, r in enumerate(rows):
        if any(c is not None and str(c).strip() != "" for c in r):
            return i
    return 0


def _read_xlsx_bytes(data: bytes) -> tuple[str, list[str], list[dict]]:
    wb = _openpyxl.load_workbook(_io.BytesIO(data), data_only=True)
    ws = wb.active
    all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
    if not all_rows:
        return ws.title, [], []
    header_idx = _detect_header_row(all_rows)
    raw = all_rows[header_idx]
    headers: list[str] = []
    seen: dict[str, int] = {}
    for c in raw:
        h = "" if c is None else str(c).strip()
        if h == "":
            headers.append(""); continue
        if h in seen:
            seen[h] += 1
            headers.append(f"{h} ({seen[h]})")
        else:
            seen[h] = 1
            headers.append(h)
    rows = []
    for r in all_rows[header_idx + 1:]:
        if all(c is None or (isinstance(c, str) and c.strip() == "") for c in r):
            continue
        rec = {}
        for i, h in enumerate(headers):
            if not h: continue
            rec[h] = r[i] if i < len(r) else None
        rows.append(rec)
    return ws.title, [h for h in headers if h], rows


def _read_csv_bytes(data: bytes) -> tuple[str, list[str], list[dict]]:
    try:
        text = data.decode("utf-8-sig", errors="replace")
    except Exception:
        text = data.decode("latin-1", errors="replace")
    reader = _csv.reader(_io.StringIO(text))
    all_rows = [list(r) for r in reader]
    if not all_rows:
        return "csv", [], []
    header_idx = _detect_header_row(all_rows)
    raw = all_rows[header_idx]
    headers: list[str] = []
    seen: dict[str, int] = {}
    for c in raw:
        h = "" if c is None else str(c).strip()
        if h == "":
            headers.append(""); continue
        if h in seen:
            seen[h] += 1
            headers.append(f"{h} ({seen[h]})")
        else:
            seen[h] = 1
            headers.append(h)
    rows = []
    for r in all_rows[header_idx + 1:]:
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        rec = {}
        for i, h in enumerate(headers):
            if not h: continue
            rec[h] = r[i] if i < len(r) else None
        rows.append(rec)
    return "csv", [h for h in headers if h], rows


def _read_uploaded(uploaded) -> tuple[str, list[str], list[dict]]:
    name = uploaded.name.lower()
    data = uploaded.getvalue()
    if name.endswith(".csv") or name.endswith(".tsv"):
        return _read_csv_bytes(data)
    return _read_xlsx_bytes(data)


def _suggest_key(headers_a, headers_b) -> str | None:
    common = [h for h in headers_a if h in set(headers_b)]
    common_lower = {h.lower(): h for h in common}
    for hint in _KEY_HINTS:
        if hint.lower() in common_lower:
            return common_lower[hint.lower()]
    return common[0] if common else None


def _row_key(rec, key_cols):
    return "\t".join(_norm_value(rec.get(c)) for c in key_cols)


def _compare_rows(rows_a, rows_b, key_cols, compare_cols):
    """Return (new_rows, mod_summary, mod_details, rem_rows). key_cols supports composite keys."""
    by_a: dict[str, dict] = {}
    by_b: dict[str, dict] = {}
    for r in rows_a:
        k = _row_key(r, key_cols)
        if k.replace("\t", "") != "":
            by_a.setdefault(k, r)
    for r in rows_b:
        k = _row_key(r, key_cols)
        if k.replace("\t", "") != "":
            by_b.setdefault(k, r)

    sa, sb = set(by_a), set(by_b)
    new_keys = sorted(sb - sa)
    rem_keys = sorted(sa - sb)
    com_keys = sorted(sa & sb)
    key_set  = set(key_cols)

    def _key_rec(k):
        parts = k.split("\t")
        return {key_cols[i]: parts[i] if i < len(parts) else "" for i in range(len(key_cols))}

    new_out = []
    for k in new_keys:
        rec = _key_rec(k)
        for c in compare_cols:
            if c in key_set: continue
            rec[c] = _display_value(by_b[k].get(c))
        new_out.append(rec)

    rem_out = []
    for k in rem_keys:
        rec = _key_rec(k)
        for c in compare_cols:
            if c in key_set: continue
            rec[c] = _display_value(by_a[k].get(c))
        rem_out.append(rec)

    mod_summary: list[dict] = []
    mod_details: list[dict] = []
    for k in com_keys:
        ra = by_a[k]; rb = by_b[k]
        key_rec = _key_rec(k)
        changed: list[str] = []
        for c in compare_cols:
            if c in key_set: continue
            va = _norm_value(ra.get(c))
            vb = _norm_value(rb.get(c))
            if va != vb:
                changed.append(c)
                mod_details.append({
                    **key_rec,
                    "Column": c,
                    "Old Value": _display_value(ra.get(c)),
                    "New Value": _display_value(rb.get(c)),
                })
        if changed:
            mod_summary.append({
                **key_rec,
                "Columns Changed": ", ".join(changed),
                "Change Count": len(changed),
            })

    return new_out, mod_summary, mod_details, rem_out


def _build_compare_excel(new_rows, mod_summary, mod_details, rem_rows, compare_cols, key_cols) -> bytes:
    wb = _openpyxl.Workbook()

    def write(name, rows, headers, color, first=False):
        ws = wb.active if first else wb.create_sheet(name)
        if first:
            ws.title = name
        ws.append(headers)
        for col, _ in enumerate(headers, 1):
            c = ws.cell(row=1, column=col)
            c.font = _Font(bold=True, color="FFFFFF")
            c.fill = _PatternFill("solid", fgColor=color)
            c.alignment = _Alignment(vertical="center")
        for r in rows:
            ws.append([r.get(h, "") for h in headers])
        for ci, h in enumerate(headers, 1):
            samples = [len(str(h))] + [min(len(str(r.get(h, ""))), 80) for r in rows[:300]]
            ml = max(samples) if samples else 12
            ws.column_dimensions[_openpyxl.utils.get_column_letter(ci)].width = min(max(12, ml + 2), 80)
        ws.freeze_panes = "A2"

    key_set = set(key_cols)
    other_cols = [c for c in compare_cols if c not in key_set]
    full_headers = list(key_cols) + other_cols
    summary_hdrs = list(key_cols) + ["Columns Changed", "Change Count"]
    details_hdrs = list(key_cols) + ["Column", "Old Value", "New Value"]

    write("New", new_rows, full_headers, "2E7D32", first=True)
    write("Modified (Summary)", mod_summary, summary_hdrs, "1565C0")
    write("Modified (Details)", mod_details, details_hdrs, "3949AB")
    write("No Longer In File", rem_rows, full_headers, "C62828")

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_spreadsheet_compare() -> None:
    st.markdown(
        """
        <div class="hub-hero">
            <h1>Spreadsheet Comparison</h1>
            <p>Upload two snapshots of the same spreadsheet (the older version and the newer version).
            The tool matches rows by a key column you pick, reports new, removed, and changed rows,
            and ignores formatting-only differences like date padding, time padding, dollar-sign
            spacing, parens-vs-minus, and CSV whitespace.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.expander("How it works", expanded=False):
        st.markdown(
            "- Upload the **starting** spreadsheet (older) and the **ending** spreadsheet (newer).\n"
            "- The tool auto-detects a likely key column (Beam Number, Case Number, Account ID, etc.) "
            "and lets you adjust the selection.\n"
            "- If one column isn't unique per row (for example, a case has multiple garnishments), "
            "you can add more columns to make a composite key.\n"
            "- A row is flagged **Modified** only when one or more cell values differ after "
            "normalization (dates → `MM/DD/YYYY`, times → `HH:MM:SS AM/PM`, money → `$0.00` form, "
            "whitespace and quoting collapsed, control chars stripped).\n"
            "- Supports `.xlsx` and `.csv`. The first sheet of an `.xlsx` is used. Blank rows above "
            "the header are detected and skipped."
        )

    st.markdown('<div class="section-label">1. Upload the starting spreadsheet</div>', unsafe_allow_html=True)
    start_upload = st.file_uploader(
        "Starting spreadsheet (older snapshot)",
        type=["xlsx", "xls", "csv"],
        accept_multiple_files=False,
        key="sc_start_file",
    )

    st.markdown('<div class="section-label">2. Upload the ending spreadsheet</div>', unsafe_allow_html=True)
    end_upload = st.file_uploader(
        "Ending spreadsheet (newer snapshot)",
        type=["xlsx", "xls", "csv"],
        accept_multiple_files=False,
        key="sc_end_file",
    )

    if not (start_upload and end_upload):
        st.info("Upload both spreadsheets to continue.")
        return

    try:
        _sheet_a, headers_a, rows_a = _read_uploaded(start_upload)
    except Exception as exc:
        st.error(f"Could not read the starting spreadsheet: {exc}")
        return
    try:
        _sheet_b, headers_b, rows_b = _read_uploaded(end_upload)
    except Exception as exc:
        st.error(f"Could not read the ending spreadsheet: {exc}")
        return

    set_a, set_b = set(headers_a), set(headers_b)
    common_cols = [h for h in headers_a if h in set_b]
    only_a = [h for h in headers_a if h not in set_b]
    only_b = [h for h in headers_b if h not in set_a]

    st.markdown(
        f"""
        <div class="metric-strip">
            <div class="metric-box"><strong>{len(rows_a)}</strong><span>rows in starting file</span></div>
            <div class="metric-box"><strong>{len(rows_b)}</strong><span>rows in ending file</span></div>
            <div class="metric-box"><strong>{len(common_cols)}</strong><span>shared columns</span></div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if not common_cols:
        st.error(
            "The two spreadsheets have no shared columns, so rows can't be matched. "
            "Make sure both files have the same column headers."
        )
        if only_a:
            st.write(f"**Only in starting file:** {', '.join(only_a)}")
        if only_b:
            st.write(f"**Only in ending file:** {', '.join(only_b)}")
        return

    if only_a or only_b:
        with st.expander("Column differences (header-level)", expanded=False):
            if only_a:
                st.write(f"**Columns only in starting file:** {', '.join(only_a)}")
            if only_b:
                st.write(f"**Columns only in ending file:** {', '.join(only_b)}")
            st.caption("These columns are ignored when comparing rows. Only shared columns are compared.")

    suggested = _suggest_key(headers_a, headers_b)
    default_keys = [suggested] if suggested in common_cols else ([common_cols[0]] if common_cols else [])
    key_cols = st.multiselect(
        "Key column(s) — pick one or more to uniquely identify a row",
        options=common_cols,
        default=default_keys,
        help="Auto-detected key is selected. If one column isn't unique (e.g. a case has multiple garnishments), add more columns to make the key unique per row.",
        key="sc_key_cols",
    )
    if not key_cols:
        st.warning("Pick at least one key column.")
        return

    # Duplicate-key check on the chosen key
    keys_a_all = [_row_key(r, key_cols) for r in rows_a]
    keys_b_all = [_row_key(r, key_cols) for r in rows_b]
    dup_a = sum(1 for c in _Counter(keys_a_all).values() if c > 1)
    dup_b = sum(1 for c in _Counter(keys_b_all).values() if c > 1)
    if dup_a or dup_b:
        st.warning(
            f"Heads up: with the key column(s) you picked, **{dup_a}** duplicate key(s) exist in the "
            f"starting file and **{dup_b}** in the ending file. Only the FIRST row for each duplicate "
            "is compared. Add more columns to the key to make rows unique."
        )

    if st.button("Compare spreadsheets", type="primary", use_container_width=True):
        progress = st.progress(0, text="Comparing rows...")
        new_rows, mod_summary, mod_details, rem_rows = _compare_rows(
            rows_a, rows_b, key_cols, common_cols
        )
        progress.progress(80, text="Building Excel report...")
        excel_bytes = _build_compare_excel(
            new_rows, mod_summary, mod_details, rem_rows, common_cols, key_cols
        )
        progress.progress(100, text="Done")

        st.session_state["sc_new"]      = new_rows
        st.session_state["sc_summary"]  = mod_summary
        st.session_state["sc_details"]  = mod_details
        st.session_state["sc_rem"]      = rem_rows
        st.session_state["sc_excel"]    = excel_bytes
        st.session_state["sc_keys"]     = key_cols
        common_count = len(set(keys_a_all) & set(keys_b_all))
        st.session_state["sc_unchanged"] = max(0, common_count - len(mod_summary))

    if "sc_excel" in st.session_state:
        new_rows    = st.session_state["sc_new"]
        mod_summary = st.session_state["sc_summary"]
        mod_details = st.session_state["sc_details"]
        rem_rows    = st.session_state["sc_rem"]
        unchanged   = st.session_state.get("sc_unchanged", 0)

        st.markdown(
            f"""
            <div class="metric-strip">
                <div class="metric-box"><strong>{len(new_rows)}</strong><span>new rows</span></div>
                <div class="metric-box"><strong>{len(mod_summary)}</strong><span>modified rows ({len(mod_details)} cell changes)</span></div>
                <div class="metric-box"><strong>{len(rem_rows)}</strong><span>no longer in file</span></div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.caption(f"{unchanged} row(s) had no real changes after normalization (formatting-only differences ignored).")

        st.download_button(
            "Download Excel report",
            data=st.session_state["sc_excel"],
            file_name="Spreadsheet Comparison.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )

        tab_new, tab_sum, tab_det, tab_rem = st.tabs(
            ["New", "Modified (Summary)", "Modified (Details)", "No Longer In File"]
        )
        with tab_new:
            if new_rows:
                st.dataframe(new_rows, use_container_width=True, hide_index=True)
            else:
                st.info("No new rows.")
        with tab_sum:
            if mod_summary:
                st.dataframe(mod_summary, use_container_width=True, hide_index=True)
            else:
                st.info("No modified rows.")
        with tab_det:
            if mod_details:
                st.dataframe(mod_details, use_container_width=True, hide_index=True)
            else:
                st.info("No cell-level changes.")
        with tab_rem:
            if rem_rows:
                st.dataframe(rem_rows, use_container_width=True, hide_index=True)
            else:
                st.info("No removed rows.")


def render_beam_pdf_splitter() -> None:
    st.markdown(
        """
        <div class="hub-hero">
            <h1>Beam PDF Splitter</h1>
            <p>Upload a PDF or a folder of PDFs, review every two-page chunk, fix any missing beam numbers, and download a clean ZIP of renamed files.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.expander("Processing rules", expanded=True):
        col1, col2 = st.columns(2)
        with col1:
            include_odd_final_page = st.checkbox("Include odd final page as a single-page PDF", value=False)
        with col2:
            use_fallback_names = st.checkbox("Use fallback name if beam number is not found", value=False)

    st.markdown('<div class="section-label">Upload PDFs</div>', unsafe_allow_html=True)
    upload_tabs = st.tabs(["Single or multiple PDF files", "Folder of PDFs"])
    with upload_tabs[0]:
        file_uploads = st.file_uploader(
            "Choose one PDF or several PDFs",
            type=["pdf"],
            accept_multiple_files=True,
            help="Use this when you have one PDF or a few PDFs selected manually.",
            key="beam_pdf_file_uploads",
        )
    with upload_tabs[1]:
        folder_uploads = st.file_uploader(
            "Choose a folder",
            type=["pdf"],
            accept_multiple_files="directory",
            help="Use this when you want to upload all PDFs from a folder.",
            key="beam_pdf_folder_uploads",
        )

    uploaded_files = list(file_uploads or []) + list(folder_uploads or [])
    upload_signature = uploaded_files_signature(uploaded_files)
    if st.session_state.get("beam_upload_signature") != upload_signature:
        st.session_state["beam_upload_signature"] = upload_signature
        st.session_state["beam_zip_bytes"] = None
        st.session_state["beam_result_rows"] = []
        st.session_state["beam_result_logs"] = []

    if not uploaded_files:
        left, right = st.columns(2)
        with left:
            st.markdown(
                """
                <div class="tool-card">
                    <h3>What it does</h3>
                    <p>Splits page pairs like 1-2, 3-4, 5-6 and names each output from the beam number at the bottom of the second page.</p>
                </div>
                """,
                unsafe_allow_html=True,
            )
        with right:
            st.markdown(
                """
                <div class="tool-card">
                    <h3>Review before download</h3>
                    <p>If OCR misses a beam number, edit the filename in the review table before creating the ZIP.</p>
                </div>
                """,
                unsafe_allow_html=True,
            )
        return

    expected_pairs = 0
    for uploaded_file in uploaded_files:
        try:
            with fitz.open(stream=uploaded_file.getvalue(), filetype="pdf") as doc:
                expected_pairs += doc.page_count // 2
        except Exception:
            pass

    st.markdown(
        f"""
        <div class="metric-strip">
            <div class="metric-box"><strong>{len(uploaded_files)}</strong><span>PDF upload(s)</span></div>
            <div class="metric-box"><strong>{expected_pairs}</strong><span>two-page output file(s)</span></div>
            <div class="metric-box"><strong>ZIP</strong><span>renamed download</span></div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.caption("After upload, click the button below. The app will create one renamed PDF for each two-page pair.")

    if st.button("Split and Rename PDFs", type="primary", use_container_width=True):
        progress = st.progress(0, text="Splitting PDFs...")
        try:
            zip_bytes, result_rows, logs = split_and_rename_pdfs(uploaded_files, include_odd_final_page, use_fallback_names)
            st.session_state["beam_zip_bytes"] = zip_bytes
            st.session_state["beam_result_rows"] = result_rows
            st.session_state["beam_result_logs"] = logs
            progress.progress(100, text="Finished")
        except Exception as exc:
            st.session_state["beam_zip_bytes"] = None
            st.session_state["beam_result_rows"] = []
            st.session_state["beam_result_logs"] = [f"Processing failed: {exc}"]
            progress.progress(100, text="Failed")

    result_rows = st.session_state.get("beam_result_rows", [])
    logs = st.session_state.get("beam_result_logs", [])

    if logs:
        with st.expander("Processing log", expanded=True):
            for log in logs:
                st.write(log)
        if not st.session_state.get("beam_zip_bytes"):
            st.error("No download was created. Check the processing log above.")

    if result_rows:
        st.subheader("Created Files")
        st.dataframe(result_rows, use_container_width=True, hide_index=True)

    if st.session_state.get("beam_zip_bytes"):
        st.success("Your split and renamed PDFs are ready.")
        st.download_button(
            "Download Beam Split Results",
            data=st.session_state["beam_zip_bytes"],
            file_name="beam_split_results.zip",
            mime="application/zip",
            type="primary",
            use_container_width=True,
        )


def render_home(tools: list[ToolDefinition]) -> None:
    st.markdown(
        """
        <div class="hub-hero">
            <h1>Resolution Law Tools</h1>
            <p>A clean web toolbox for PDF, document, and office workflows. Start with Beam PDF Splitter, then add more tools as the team needs them.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.subheader("Available Tools")
    columns = st.columns(2)
    for index, tool in enumerate(tools):
        with columns[index % 2]:
            st.markdown(
                f"""
                <div class="tool-card clickable">
                    <h3>{tool.name}</h3>
                    <p>{tool.description}</p>
                    <p class="small-muted">{tool.category} tool</p>
                </div>
                """,
                unsafe_allow_html=True,
            )
            if st.button(tool.name, key=f"open_{tool.tool_id}", use_container_width=True):
                st.session_state["active_tool_label"] = tool.name
                st.rerun()

    st.info("Future tools can be added by adding another render function and ToolDefinition in app.py.")


def render_future_tools_guide() -> None:
    st.markdown(
        """
        <div class="hub-hero">
            <h1>Add Future Tools</h1>
            <p>Resolution Law Tools is set up as a toolbox. Add another workflow by creating a render function and registering it in the tools list.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.write("Add a new function like this:")
    st.code(
        '''def render_new_tool() -> None:
    st.markdown("""
    <div class="hub-hero">
        <h1>New Tool</h1>
        <p>Short description of what this tool does.</p>
    </div>
    """, unsafe_allow_html=True)
    st.write("Build the workflow here.")
''',
        language="python",
    )

    st.write("Then add a registry entry:")
    st.code(
        '''ToolDefinition(
    tool_id="new-tool",
    name="New Tool",
    category="Documents",
    description="Short plain-English description.",
    render=render_new_tool,
)''',
        language="python",
    )


def get_tools() -> list[ToolDefinition]:
    return [
        ToolDefinition(
            tool_id="beam-pdf-splitter",
            name="Beam PDF Splitter",
            category="PDF",
            description="Split PDFs into two-page chunks and name each output from the beam number in the footer.",
            render=render_beam_pdf_splitter,
        ),
        ToolDefinition(
            tool_id="spreadsheet-compare",
            name="Spreadsheet Comparison",
            category="Documents",
            description="Compare two snapshots of a spreadsheet (.xlsx or .csv). Match rows by a key column; report new, removed, and changed rows; ignore formatting-only differences.",
            render=render_spreadsheet_compare,
        ),
    ]


def main() -> None:
    st.set_page_config(page_title=APP_TITLE, layout="wide")
    inject_css()

    tools = get_tools()
    menu_options = ["Home"] + [tool.name for tool in tools] + ["Add Future Tools"]
    active_label = st.session_state.get("active_tool_label", "Home")
    if active_label not in menu_options:
        active_label = "Home"

    with st.sidebar:
        st.title(APP_TITLE)
        selected_label = st.radio(
            "Tools",
            menu_options,
            index=menu_options.index(active_label),
            label_visibility="collapsed",
        )
        st.session_state["active_tool_label"] = selected_label
        st.divider()
        st.caption("Upload files, process them in the browser app, then download results.")

    if selected_label == "Home":
        render_home(tools)
        return

    if selected_label == "Add Future Tools":
        render_future_tools_guide()
        return

    selected_tool = next(tool for tool in tools if tool.name == selected_label)
    selected_tool.render()


if __name__ == "__main__":
    main()
